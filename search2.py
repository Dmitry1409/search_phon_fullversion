import os, openpyxl, pickle, csv, shutil, random
from openpyxl import Workbook



BASA_FILE_NAME = "BASA"
DIRECTORY_AVITO_PARSER = 'архив ави парсер'
DIRECTORY_REPORT_ARCH = 'REPORTS/архив'
DIRECTORY_NEW_REPORT = 'REPORTS/new'
FEEDBACK_FILE_NAME = 'feedback_list'
RESERV_COPY_FEEDBACK_FILE_NAME = 'resrv_feed_back'

def main():
	basa = init()
	response = None
	while response != '0':
		print('''
			1 - Проверить файл
			2 - Скрипт выгрузки
			3 - Поиск по номеру
			4 - feedback
			5 - Валидация имен
			6 - 
			10 - Проверка категорий
			0 - Выйти''')
		response = input("Введите: ")

		if response == '1':
			res1(basa)
		if response == '2':
			res2(basa)


		if response == '3':
			try:
				print(basa[phone_format()])
			except:
				print('Не найдено')

		if response == '4':
			add_feedback()

		if response == '5':
			import name2
			name2.main()
		

def res1(basa):
	file_name = input('Имя файла: ') + '.xlsx'
	if chek_file(file_name):
		file = read_exel_in_dict(file_name)
		print('В файле ', len(file), " номеров.")
		clean_dt = {}
		defolt = {}
		for k in file:
			fl = True
			for kb in basa:
				if kb == k:
					fl = False
					break
			if fl:
				clean_dt[k] = file[k]
			else:
				defolt[k] = file[k]
		if clean_dt:
			print('Найдено ', len(defolt), ' совподений из базы.')
			request = 'Записать '+ str(len(clean_dt)) + ' номеров в базу? '
			if input(request) == 'y':
				add_atribute(clean_dt)
				basa.update(clean_dt)
				save_pickle(BASA_FILE_NAME, basa)
				move_to(file_name, DIRECTORY_AVITO_PARSER)
	else:
		print('File ', file_name,' not found.')

def res2(basa):
	targ = {}
	for k in basa:
		if basa[k]['Отправлено'] == 0:
			targ[k] = basa[k]
	print(len(targ))
	categ = separate_categor(basa)
	print(len(categ))
	targ2 = search_data_deliver(categ, 20200401)
	print(len(targ2))
	count = 0
	for k in targ2:
		if k in targ:
			count +=1
	print(count, ' счет')
	targ.update(targ2)
	shif_list = list(targ.keys())
	random.shuffle(shif_list)
	rec_dt = {}
	for k in shif_list:
		rec_dt[k] = targ[k]
	writin_new_exele('shuf.xlsx', rec_dt)



def separate_categor(basa):
	def categ(categ, dt):
		map_dt = {}
		for k in dt:
			# if '/' in dt[k][categ]:

			# else:	
			if map_dt.get(dt[k][categ]):
				map_dt[dt[k][categ]] += 1
			else:
				map_dt[dt[k][categ]] = 1
		return map_dt

	def print_key(map_dt, categ):
		key_list = list(map_dt.keys())
		i = 1
		for j in key_list:
			print(i, j, map_dt[j])
			i +=1
		targ = {}
		resp = input('Введите: ')
		if resp == 'x':
			pass
			# for k in basa:
			# 	if 

		for k in basa:
			if basa[k][categ] == key_list[int(resp)-1]:
				targ[k] = basa[k]


		return targ

	
	map_dt = categ('Подкатегория', basa)
	targ = print_key(map_dt, 'Подкатегория')
	
	return targ

	# формат red_line (int: 20200914)
def search_data_deliver(dt, red_line):
	def spliting(string):
		data = string.split(' ')[:1]
		data = data[0].split('.')
		data.reverse()
		data = ''.join(data)
		return data
	tar = {}
	for k in dt:
		if len(dt[k]['Дата отправки'])>1:
			cup = []
			for s in dt[k]['Дата отправки']:
				cup.append(spliting(s))
			if int(max(cup)) <= red_line:
				tar[k] = dt[k]
		elif len(dt[k]['Дата отправки']) == 1:
			if int(spliting(dt[k]['Дата отправки'][0])) <= red_line:
				tar[k] = dt[k]
		else:
			tar[k] = dt[k]
	return tar

def move_to(file_name, to_directory, from_directory=None):
	if from_directory:
		os.rename(from_directory+'/'+file_name, to_directory +'/' + file_name)
		print('Move from: ',  from_directory+'/'+file_name)
		print('Move to: ', to_directory+'/'+file_name)
	else:
		os.rename(file_name, to_directory +'/' + file_name)
		print('Move from: ',  file_name)
		print('Move to: ', to_directory+'/'+file_name)

def save_pickle(file_name, file):
	if chek_file(file_name):
		with open(file_name, 'rb+') as D:
			pickle.dump(file, D)
	else:
		with open(file_name, 'wb+') as D:
			pickle.dump(file, D)

	print('Pickle dump ', os.getcwd()+'\\'+ file_name)

def is_validate_sity(sity):
	SITY_VAL = ["Нижний Новгород", "Дзержинск", "Бор", "Кстово" ,
				 "Павлово" , "Богородск" , "Городец", "Балахна", 
				 "Семенов", "Заволжье", "Чкаловск", "Володарск", "Ворсма",
				  "Ковернино", "Большое Козино", "Лукино", "Линда"]
	for s in SITY_VAL:
		if s == sity:
			return True
	return False

def writin_new_exele(path, rec_dt, writin_keys = True):
	
	def get_keys(dt):
		l_key = ['Телефон']
		for i in dt:
			for i2 in dt[i]:
				l_key.append(i2)
			break
		return l_key

	def dict_in_rec_list(rec_dt, rec_list):
		l_key = rec_list[0]
		for k in rec_dt:
			l_val = [k]
			for i in l_key[1:]:
				if i == 'Дата отправки' or i == 'Текст сообщения':
					if rec_dt[k][i] == []:
						l_val.append(None)
					else:
						l_val.append(', '.join(rec_dt[k][i]))
					continue
				l_val.append(rec_dt[k].get(i))
			rec_list.append(l_val)

		return rec_list

	def save_exele(path, list_val):
		wb = Workbook()
		ws = wb.active
		for row in list_val:
			ws.append(row)
		wb.save(path)

	print('recording: ', path)
	l_key = get_keys(rec_dt)
	l_val_main = [l_key]
	l_val_main = dict_in_rec_list(rec_dt, l_val_main)
	
	if writin_keys == False:
		return save_exele(path, l_val_main[1:])
	
	save_exele(path, l_val_main)

def search_bulding_person(s_dt):
	main_w = ['ремонт', 'строит', 'отдел', 'стройка', 'плиточ', 'электр', 'сантехн', 'плотник',
			 'грузчик', 'лестниц', 'фасад', 'фундамент', 'каменщ','монтаж',
			  "печник", "кровля", "металлоконструкции", "забор", "плотн", "рабоч", "штукат"]

	dub_w = ['ремонт', 'отдел', 'строит', 'клей', 'штукату', 'каркас', 'плит', "электр",
			"работ", "услуг", "монтаж", "установ", "уклад", "утепл", "дизай", "проект", "мастер",
			"качес", "сантех", "изол", "шпак", "бан"]

	obj_w = ['офис', 'дом', 'квар', 'рабо', 'бригад', 'услуг', 'капит', 'плит',
			 'обо', 'гипс', 'внутр', 'помещ', 'ванн', 'туал', 'ключ', 'коттедж',
			 "помещ", "бан", 'унив', "сруб", "интер", "фундам", "кров", "фасад",
			 "отоп", "стен"]

	t_dt = {}
	for n in s_dt:
		title = s_dt[n].get('Заголовок')
		if title:
			title = title.strip().lower()
			if ' ' not in title:
				for w in main_w:
					if w in title:
						t_dt[n] = s_dt[n]
						break
			else:
				fd = 0
				for w in dub_w:
					if w in title:
						fd +=1
				if fd > 1:
					t_dt[n] = s_dt[n]
					continue
				if fd == 1:
					fo = False
					for ow in obj_w:
						if ow in title:
							t_dt[n] = s_dt[n]
							fo = True
							break
	return t_dt

def sort_dic_for_rec(dt):
	map_dt = {}
	l_dict = []
	for k in dt:
		map_dt[k] = False
	for k in dt:
		if map_dt[k]:
			continue
		map_dt[k] = True
		prov = {}
		prov[k] = dt[k]
		for k2 in dt:
			if map_dt[k2]:
				continue
			# if prov[k]['Категории'] == dt[k2]['Категории']:
			if prov[k]['Город'] == dt[k2]['Город']:
				prov[k2] = dt[k2]
				map_dt[k2] = True
		l_dict.append(prov)
	l_dict.sort(key=len)
	# l_dict.reverse()
	return l_dict

def is_valid_coord(coord):
	noth_west = (56.794862261400546, 42.286376953125)
	noth_east = (56.96893619436121, 44.57153320312501)
	south_east = (56.105746831832064, 44.5330810546875)
	south_west = (55.54417295022067, 43.4619140625)
	if coord[0] < noth_west[0] and coord[1] > noth_west[1]:
		if coord[0] < noth_east[0] and coord[1] < noth_east[1]:
			if coord[0] > south_west[0] and coord[1] > south_west[1]:
				if coord[0] > south_east[0] and coord[1] < south_east[1]:
					return True
	return False

def add_atribute(dt):
		for n in dt:
			dt[n]['Отправлено'] = 0
			dt[n]['Доставлено'] = 0
			dt[n]['Дата отправки'] = []
			dt[n]['Текст сообщения'] = []

def init():

	global_report = {'Отправлено': 0, 'Доставлено': 0}

	def greate_report(csv_list):
		rep_dt = {'send': 0, 'delived': 0, 'errors':0, 'notFound': 0}		
	
		for row in range(1, len(csv_list)):
			list_data = csv_list[row][0].split(';')
			if basa.get(list_data[1]):
				if list_data[4] != '0':
					rep_dt['send'] +=1
					rep_dt['delived'] += int(list_data[5])
					basa[list_data[1]]['Отправлено'] += int(list_data[4])
					basa[list_data[1]]['Доставлено'] += int(list_data[5])
					if basa[list_data[1]]['Дата отправки']:
						basa[list_data[1]]['Дата отправки'].append(list_data[3][1:-1])
						basa[list_data[1]]['Текст сообщения'].append(list_data[7][1:-1])
					else:
						basa[list_data[1]]['Дата отправки'] = [list_data[3][1:-1]]
						basa[list_data[1]]['Текст сообщения'] = [list_data[7][1:-1]]
				else:
					rep_dt['errors'] += 1
			else:
				rep_dt['notFound'] +=1

		return rep_dt

	if chek_file(BASA_FILE_NAME):
		with open(BASA_FILE_NAME, 'rb') as D:
			basa = pickle.load(D)
		print('База загружена. ', len(basa), ' номеров.')
	else:
		if input('База не найдена, создать?: ') == 'y':
			print("Создается новая база")
			basa = {}
			path_list_xl = get_path(DIRECTORY_AVITO_PARSER)
			for p in path_list_xl:
				dt = read_exel_in_dict(p)
				basa.update(dt)
			add_atribute(basa)
			path_report = get_path(DIRECTORY_REPORT_ARCH)

			rep_dt = {'send': 0, 'delived': 0, 'errors':0, 'notFound': 0}

			for p in path_report:
				csv_list = read_csv(p)
				
				
				# Там есть два файла которие отличаются от все остальных
				# для них своя реализация
				if len(csv_list[2][0].split(';'))<2:
					i = 0
					for row in range(1, int(len(csv_list[1:])/4)+1):
						list_data = csv_list[row+i][0].split(';')
						if basa.get(list_data[1]):
							if list_data[4] != '0':
								rep_dt['send'] +=1
								rep_dt['delived'] += int(list_data[5])
								basa[list_data[1]]['Отправлено'] += int(list_data[4])
								basa[list_data[1]]['Доставлено'] += int(list_data[5])
								if basa[list_data[1]]['Дата отправки']:
									basa[list_data[1]]['Дата отправки'].append(list_data[3][1:-1])
									basa[list_data[1]]['Текст сообщения'].append(list_data[7][1:] + csv_list[row+i+1][0] + csv_list[row+i+2][0] + csv_list[row+i+3][0])
								else:
									basa[list_data[1]]['Дата отправки'] = [list_data[3][1:-1]]
									basa[list_data[1]]['Текст сообщения'] = [list_data[7][1:] + csv_list[row+i+1][0] + csv_list[row+i+2][0] + csv_list[row+i+3][0]]
						else:
							rep_dt['notFound'] +=1
						i+= 3
					
				else:
					report = greate_report(csv_list)
					for k in rep_dt:
						rep_dt[k] += report[k]
			print(rep_dt)
			save_pickle(BASA_FILE_NAME, basa)
			print('Создалась новая база. ', len(basa), ' номеров.')

	# проверяем новые отчеты
	path_report = os.listdir(DIRECTORY_NEW_REPORT)
	if path_report:
		print(path_report)
		if input('Выполнить?: ') == 'y':
			rep_dt = {'send': 0, 'delived': 0, 'errors':0, 'notFound': 0}
			for p in path_report:
				csv_list = read_csv('REPORTS/new/'+p)
				report = greate_report(csv_list)
				for k in rep_dt:
					rep_dt[k] += report[k]
				move_to(p, DIRECTORY_REPORT_ARCH, from_directory=DIRECTORY_NEW_REPORT)
			print(rep_dt)
			save_pickle(BASA_FILE_NAME, basa)				

	for k in basa:
		global_report['Отправлено'] += basa[k]['Отправлено']
		global_report['Доставлено'] += basa[k]['Доставлено']

	print('Доставлено: ', global_report['Доставлено'])
	print('Отправлено: ', global_report['Отправлено'])

	return basa

def read_csv(path):
	print('Load report: ', path)
	csv_list = []
	with open(path, newline = '', encoding= 'utf-8') as f:
		reader = csv.reader(f)
		for row in reader:
			csv_list.append(row)
	return csv_list

def read_exel_in_dict(path):
	print("load: ", path)
	CHAR = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j' , 'k', 'l', 'm', 'n',
			'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'aa', 'ab',
			'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an',
			'ao', 'ap', 'aq', 'ar', 'as']
	exel = openpyxl.load_workbook(path)
	sheet = exel.active
	cols = sheet.max_column
	rows = sheet.max_row
	l_key = []	
	for s in range(0, cols):
		k = CHAR[s] + '1'
		l_key.append(sheet[k].value)
	dt = {}	
	for r in range(2, rows +1):
		phone = sheet[CHAR[0] + str(r)].value 
		dt[phone] = {}
		for i in range(1, len(l_key)):
			dt[phone][l_key[i]] = sheet[CHAR[i] + str(r)].value
	return dt

def get_path(directory = None, type_fl = None):
	if type_fl:
		pint('raliase func')
	else:
		L = os.listdir(directory)
		path_list = []
		for e in L:
			path_list.append(directory+'/'+e)
		return path_list

def chek_file(path):
	l = path.split("/")
	if len(l)>1:
		dir_ = "/".join(l[:-1])
	else:
		dir_ = None
	list_dir = os.listdir(dir_)
	flag = False
	for f in list_dir:
		if f == l[-1]:
			flag = True
			break
	return flag

def phone_format():
	phone = input('Телефон +7ххх ххх хх хх: ')
	if phone[:2] =='+7' and len(phone) == 12:
		return phone
	elif phone[0] == '8' and len(phone) == 11:
		return '+7' + phone[1:]
	elif phone[0] == '9' and len(phone) == 10:
		return '+7' + phone
	else:
		print('Некорректный номер ', phone)

def init_feedback(dt = None):
	fl = chek_file(FEEDBACK_FILE_NAME)
	if dt:
		if fl:
			with open(FEEDBACK_FILE_NAME, 'rb') as f:
				pic = pickle.load(f)				
			pic.update(dt)
			save_pickle(FEEDBACK_FILE_NAME, pic)
		else:
			save_pickle(FEEDBACK_FILE_NAME, dt)
	else:
		if fl:	
			with open(FEEDBACK_FILE_NAME, 'rb') as f:
				dt = pickle.load(f)
			return dt
		else:
			return {}

def add_feedback():

	changed = False

	def search_or_delet(dt, phone, d = False):
		l_key = list(dt)
		fl = False
		for k in l_key:
			l_k2 = list(dt[k])
			for k2 in l_k2:
				if k2 == phone:
					if d:
						del dt[k][k2]
					fl = True
					break
			if fl:
				break
		if d:
			return dt, k
		else:		
			return fl, k

	def resp_key(keys, str_ob):
		i = 1
		print('Разделы')
		for k in keys:
			print(i, " - ", k)
			i += 1
		print()
		print('a - добавить раздел')
		if str_ob == 'выйти.':
			print('v - показать раздел')
			print('d - удалить номер')
		if str_ob == "выйти.":
			print('s - сохраниться')
		print('x - ', str_ob)
		print()
		resp = input('Ввод: ')
		return resp

	dt = init_feedback()
	resp = None
	while resp != 'x':
		keys = list(dt)
		keys.sort()
		resp = resp_key(keys, 'выйти.')
		if resp =="s":
			init_feedback(dt)
			changed = True
		if resp =="a":
			n = input("Наименование: ")
			dt[n] = {}
		if resp =='v':
			r = input('Номер или all: ')
			if r.isdigit():
				if int(r) in range(0, len(keys)+1) and r != '0':
					l_k2 = list(dt[keys[int(r)-1]])
					for k2 in l_k2:
						print(k2, dt[keys[int(r)-1]][k2])
				else:
					print('Нет такого раздела')
			if r == 'all':
				print(dt)
		if resp == 'd':
			phone = phone_format()
			fl = search_or_delet(dt, phone)
			if fl:
				c = search_or_delet(dt, phone, d= True)
				dt = c[0]
				print('Номер ', phone, ' удален.')
			else:
				print('Номер ', phone, ' не найден.')

		if resp =='x':
			if changed:
				shutil.copy2(FEEDBACK_FILE_NAME, RESERV_COPY_FEEDBACK_FILE_NAME)
				print('Резервное копирование выполнено.')
				changed = False

		if resp.isdigit():
			if int(resp) in range(0, len(keys)+1) and resp != '0':
				phone = phone_format()
				fl = search_or_delet(dt, phone)
				if fl[0] == False:
					dt[keys[int(resp)-1]][phone] ={}
					resp2 = None
				else:
					print('Номер ', phone, ' найден в ', fl[1]) 
					resp2 = 'x'
				name_catalog = keys[int(resp) -1]
				print(name_catalog)
				phone_l = list(dt[name_catalog])
				keys2 = []
				for n in phone_l:
					for k in dt[name_catalog][n]:
						if k not in keys2:
							keys2.append(k)
				keys2.sort()
				while resp2 != 'x':
					print(phone, '  ', dt[name_catalog][phone])																		
					print()
					resp2 = resp_key(keys2, 'назад')
					if resp2 =="a":
						n = input("Наименование: ")
						v = input('Значение: ')
						dt[name_catalog][phone][n] = v 
					if resp2.isdigit():
						if int(resp2) in range(0, len(keys2)+1) and resp2 != '0':
							value = input("Значение: ")
							dt[name_catalog][phone][keys2[int(resp2)-1]] = value

if __name__ == '__main__':
	main()