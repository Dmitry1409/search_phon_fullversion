import openpyxl
from openpyxl import Workbook
import os
import pickle
import csv

def validate_celling(dt):
	resp = {}
	for k in dt:
		if 'натяжн' not in dt[k]['Заголовок'].lower():
			resp[k] = dt[k]
	return resp

def validate_sity(dt):
	SITY_VAL = ["Нижний Новгород", "Дзержинск", "Бор", "Кстово" , "Павлово" , "Богородск" , "Городец", 
				"Балахна", "Семенов", "Заволжье", "Чкаловск", "Володарск", "Ворсма", "Ковернино", "Большое Козино",
				 "Лукино", "Линда"]
	resp = {}
	for k in dt:
		for s in SITY_VAL:
			if s.lower() == dt[k]['Город'].lower():
				resp[k] = dt[k]
				break
	return resp

def add_feedback():
	
	def search_or_delet(dt, phone, d = False):
		l_key = list(dt)
		fl = False
		for k in l_key:
			l_k2 = list(dt[k])
			for k2 in l_k2:
				if k2 == phone:
					if d:
						del dt[k][k2]
					fl = True
					break
			if fl:
				break
		if d:
			return dt, k
		else:		
			return fl, k

	def resp_key(keys, str_ob):
		i = 1
		print('Разделы')
		for k in keys:
			print(i, " - ", k)
			i += 1
		print()
		print('a - добавить раздел')
		if str_ob == 'выйти.':
			print('v - показать раздел')
			print('d - удалить номер')
		if str_ob == "выйти.":
			print('s - сохраниться')
		print('x - ', str_ob)
		print()
		resp = input('Ввод: ')
		return resp

	dt = init_feedback()
	resp = None
	while resp != 'x':
		keys = list(dt)
		keys.sort()
		resp = resp_key(keys, 'выйти.')
		if resp =="s":
			init_feedback(dt)
		if resp =="a":
			n = input("Наименование: ")
			dt[n] = {}
		if resp =='v':
			r = input('Номер или all: ')
			if r.isdigit():
				if int(r) in range(0, len(keys)+1) and r != '0':
					l_k2 = list(dt[keys[int(r)-1]])
					for k2 in l_k2:
						print(k2, dt[keys[int(r)-1]][k2])
				else:
					print('Нет такого раздела')
			if r == 'all':
				print(dt)
		if resp == 'd':
			phone = phone_format()
			fl = search_or_delet(dt, phone)
			if fl:
				c = search_or_delet(dt, phone, d= True)
				dt = c[0]
				print('Номер ', phone, ' удален.')
			else:
				print('Номер ', phone, ' не найден.')

		if resp.isdigit():
			if int(resp) in range(0, len(keys)+1) and resp != '0':
				phone = phone_format()
				fl = search_or_delet(dt, phone)
				if fl[0] == False:
					dt[keys[int(resp)-1]][phone] ={}
					resp2 = None
				else:
					print('Номер ', phone, ' найден в ', fl[1]) 
					resp2 = 'x'
				name_catalog = keys[int(resp) -1]
				print(name_catalog)
				phone_l = list(dt[name_catalog])
				keys2 = []
				for n in phone_l:
					for k in dt[name_catalog][n]:
						if k not in keys2:
							keys2.append(k)
				keys2.sort()
				while resp2 != 'x':
					print(phone, '  ', dt[name_catalog][phone])																		
					print()
					resp2 = resp_key(keys2, 'назад')
					if resp2 =="a":
						n = input("Наименование: ")
						v = input('Значение: ')
						dt[name_catalog][phone][n] = v 
					if resp2.isdigit():
						if int(resp2) in range(0, len(keys2)+1) and resp2 != '0':
							value = input("Значение: ")
							dt[name_catalog][phone][keys2[int(resp2)-1]] = value

def phone_format():
	phone = input('Телефон +7ххх ххх хх хх: ')
	if phone[:2] =='+7' and len(phone) == 12:
		return phone
	elif phone[0] == '8' and len(phone) == 11:
		return '+7' + phone[1:]
	elif phone[0] == '9' and len(phone) == 10:
		return '+7' + phone
	else:
		print('Некорректный номер ', phone)

def init_feedback(dt = None):
	fl = chek_file('feedback_list')
	if dt:
		if fl:
			with open("feedback_list", 'rb') as f:
				pic = pickle.load(f)				
			pic.update(dt)
			with open('feedback_list', 'rb+') as f:
				pickle.dump(pic, f)
				print('Файл feedback обнавлен ')
		else:
			with open('feedback_list', 'wb') as f:
				pickle.dump(dt, f)
			print('Создан словарь feedback')
			print('Добавлено ', len(dt))
	else:
		if fl:	
			with open("feedback_list", 'rb') as f:
				dt = pickle.load(f)
			return dt
		else:
			return {}
		
def response10(main_dict):
	res = input("Введите имя файла: .xlsx") + ".xlsx"
	search_dt = read_exel_in_dict(res)
	search_dt = validate_category(search_dt)
	l = sort_dic_for_rec(search_dt)
	for d in l:
		k = list(d.keys())
		print (d[k[0]]['Категории'])

def response4(main_dict):
	rep_dt = {}
	for k in main_dict:
		if main_dict[k]["Отправлено"] > 0:
			rep_dt[k] = main_dict[k]
	rep_list = sort_dic_for_rec(rep_dt)
	rep_list.sort(key=len)
	rep_list.reverse()
	i  = 0
	for d in rep_list:
		dev = 0
		for k in d:
			dev += d[k]['Доставлено']
		print(d[k]['Категории'][47:], '/', d[k]['Город'], ' отпр-', len(d), 'дост-', dev)
		i += 1
		if i == 70:
			break
	sen = 0
	dev = 0
	for k in rep_dt:
		sen += rep_dt[k]['Отправлено']
		dev += rep_dt[k]['Доставлено']
	print()
	print('Отправлено всего ', sen, ' доставлено ', dev)

def response3(main_dict):
	resp = input('Введите номер 9хх ххх хх хх : ')
	resp = "+7" + resp
	ser_dt = {}
	fl = False
	for k in main_dict:
		if k == resp:
			ser_dt[k] = main_dict[k]
			print(k, main_dict[k])
			fl = True
	if fl:
		resp = input('Выгрузить?: ')
		if resp == 'y':
			k = list(ser_dt.keys())
			path = k[0] + '.xlsx'
			writin_new_exele(path, ser_dt)
			print('Complete')
			print(path)
	else:
		print(resp, ' не найден')

def response2(main_dict):
	sort_list = sort_dic_for_rec(main_dict)
	sort_list.sort(key=len)
	sort_list.reverse()
	ind = 0
	for d in sort_list:
		if ind == 50:
			break
		l = list(d.keys())
		print(d[l[0]]['Категории'], '/', d[l[0]]['Город'], ' ', len(d), ' номеров.')
		ind += 1

def response1(main_dict):
	Path = input('Введите имя файла .xlsx: ')
	Path = Path + '.xlsx'
	fl = chek_file(Path)
	if fl:
		search_dt = read_exel_in_dict(Path)
		search_dt = validate_category(search_dt)
	else:
		print('Не удалось открыть ', os.getcwd(), '/',  Path)
		search_dt = False
	if search_dt:
		print('В файле ', len(search_dt), ' номеров.')											# Ищем совпадение в словаре и создаем файлы для записи
		recdt, db = search_duble_in_dict(search_dt, main_dict)
		print('В файле ', Path, ' найдено ', db, ' совпадений из базы.')
		if recdt:
			mes = 'Записать в базу ' + str(len(recdt)) + ' номеров?: '
			response2 = input(mes)
			if response2 == 'y':
				os.chdir('провереные')
				l = list(recdt.keys())
				try:
					if recdt[l[0]]["Город"] and recdt[l[0]]['Категории']:						# Проверяем файл на наличее отрибутов "Категории и город"
						fl = True
				except:
					fl = False
				if fl:																			# Записываем по путям и создаем директории
					l_dict = sort_dic_for_rec(recdt)
					for dt in l_dict:
						k = list(dt.keys())																
						try:
							os.makedirs(dt[k[0]]['Категории'])
							print('Создалась директория ', dt[k[0]]['Категории'])
						except:
							pass
						path = dt[k[0]]['Категории'] + '/' + dt[k[0]]['Город'] + '.xlsx'
						fl = chek_file(path)
						if fl:
							rewrite_exele(path, dt)
							print(path, ' дозаписано ', len(dt), ' номеров.')
						else:
							writin_new_exele(path, dt)
							print(path, ' создан ', len(dt), ' номеров.')


				else:																			# Спрашиваем имя и записываем файл
					name = input('Введите имя файла .xlsx: ')
					name = name + '.xlsx'
					folder = 'Услуги/Предложение услуг/Старая база'
					os.chdir(folder)
					writin_new_exele(name, recdt)
					back_dir(3)
					print('Complete')
					print('Расположение: ', folder, '/', name)
				back_dir(1)
				os.rename(Path, 'архив ави парсер/' + Path)
				print('Перемещено: архив ави парсер/', Path )
				for r in recdt:																	# Обновляем словарь
					main_dict[r] = recdt[r]
				save_dict(main_dict)

def sort_rever_join(l_dt):
	l_dt.sort(key=len)
	l_dt.reverse()
	dt = {}
	for d in l_dt:
		dt.update(d)
	return dt

def save_dict(main_dt):
	with open('DICT.txt', 'rb+') as DICT_DEVELOR:
		pickle.dump(main_dt, DICT_DEVELOR)
	print('Словарь обновлен.', len(main_dt), ' номеров')

def search_bulding_person(s_dt):
	l_w = read_txt('search_word.txt')
	main_w = val_ser_word(l_w[0])
	dub_w = val_ser_word(l_w[1])
	obj_w = val_ser_word(l_w[2])

	t_dt = {}
	for n in s_dt:
		title = s_dt[n].get('Заголовок')
		if title:
			title = title.strip()
			title = title.lower()
			fm = False
			for w in main_w:
				if w == title:
					t_dt[n] = s_dt[n]
					fm = True
					break
			if fm:
				continue
			fd = 0
			for w in dub_w:
				if w in title:
					fd +=1
			if fd > 1:
				t_dt[n] = s_dt[n]
				continue
			if fd == 1:
				fo = False
				for ow in obj_w:
					if ow in title:
						t_dt[n] = s_dt[n]
						fo = True
						break
	return t_dt

def res_dir(main_dict):
	dir_ = load_path_dir('провереные')
	for i in range(1, len(dir_) + 1): 
		print(i, ' - ', dir_[i -1])
	res = input('Введите номер: ')
	res = dir_[int(res) - 1]
	res = res[11:]
	s_dt = {}
	for d in main_dict:
		if res in main_dict[d]['Категории']:
			s_dt[d] = main_dict[d]
	return s_dt

def val_ser_word(main_w):
	main_w = main_w.split(',')
	for i in range(0, len(main_w)):
		main_w[i] = main_w[i].strip()
		main_w[i] = main_w[i].lower()
	return main_w

def read_txt(file):
	f = open(file)
	list_word = []
	list_word = f.read().splitlines()
	f.close()
	return list_word

def sort_devilery(dt):
	dt2 = {}
	dt3 = {}
	for k in dt:
		if dt[k]['Доставлено'] == 0:
			dt2[k] = dt[k]
		else:
			dt3[k] = dt[k]
	dt2.update(dt3)
	return dt2

def sort_the_sity(d):
	map_dt = {}
	l_dict = []
	for k in d:
		map_dt[k] = False
	for k in d:
		if map_dt[k]:
			continue
		map_dt[k] = True
		prov = {}
		prov[k] = d[k]
		for k2 in d:
			if map_dt[k2]:
				continue
			if prov[k]['Город'] == d[k2]['Город']:
					prov[k2] = d[k2]
					map_dt[k2] = True
		l_dict.append(prov)
	d = sort_rever_join(l_dict)			
	return d

def load_path_dir(folder):
	path = os.listdir(folder)
	path_list = []
	for n in path:
		if n[-1] !='x':
			p = folder + '/' + n
			path_list.append(p)
			l2 = os.listdir(p)
			for i in l2:
				if i[-1] != 'x':
					p = folder + '/' + n + '/' + i
					path_list.append(p)
					l3 = os.listdir(p)
					for i2 in l3:
						if i2[-1] != 'x':
							p = folder + '/' + n + '/' + i + '/' +i2
							path_list.append(p)
							l4 = os.listdir(p)
							for i3 in l4:
								if i3[-1] != 'x':
									p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3
									path_list.append(p)
									l5 = os.listdir(p)
									for i4 in l5:
										if i4[-1] != 'x':
											p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 +'/' +i4
											path_list.append(p)
	return path_list

def dict_in_rec_list(rec_dt, rec_list):
	l_key = rec_list[0]
	for k in rec_dt:
		l_val = [k]
		for i in l_key[1:]:
			try:
				l_val.append(rec_dt[k][i])
			except:
				l_val.append('')
		rec_list.append(l_val)

	return rec_list

def remove_data(old_dt, new_dt):
	rec_dt = {}
	cop_dt = old_dt.copy()
	fl = False
	for i in new_dt:
		for i2 in old_dt:
			if i == i2:
				del(cop_dt[i])
				fl = True
	if fl:
		return cop_dt
	else:
		return old_dt

def get_keys(dt):
	l_key = ['Телефон']
	for i in dt:
		for i2 in dt[i]:
			l_key.append(i2)
		break
	return l_key

def save_exele(path, list_val):
	wb = Workbook()
	ws = wb.active
	for row in list_val:
		ws.append(row)
	wb.save(path)

def sort_dic_for_rec(dt):
	map_dt = {}
	l_dict = []
	for k in dt:
		map_dt[k] = False
	for k in dt:
		if map_dt[k]:
			continue
		map_dt[k] = True
		prov = {}
		prov[k] = dt[k]
		for k2 in dt:
			if map_dt[k2]:
				continue
			if prov[k]['Категории'] == dt[k2]['Категории']:
				if prov[k]['Город'] == dt[k2]['Город']:
					prov[k2] = dt[k2]
					map_dt[k2] = True
		l_dict.append(prov)
	l_dict.sort(key=len)
	l_dict.reverse()
	return l_dict

def value_assig(dt):
	for kd in dt:
		dt[kd]["Отправлено"] = 0
		dt[kd]["Доставлено"] = 0
		dt[kd]["Дата отправки"] = ' '
	return dt

def extrude_csv(csv):
	l_csv = []
	for r in csv:
		n = r[0].split(';')
		l_n = [n[1], n[4], n[5]]
		l_csv.append(l_n)
	return l_csv[1:]

def read_csv(path):
	csv_list = []
	with open(path, newline = '', encoding= 'utf-8') as f:
		reader = csv.reader(f)
		for row in reader:
			csv_list.append(row)
	return csv_list

def init_report(main_dt):
	p_l = os.listdir('REPORTS/new')
	if p_l:
		mess = 'Найдено ' + str(len(p_l)) + ' файлов отчета, инициализировать?: '
		res = input(mess)
		if res == 'y':																						# Собираем данные отчетов
			fl3 = False
			report_dt = {}
			for p in p_l:
				n = 0
				p = 'REPORTS/new/' + p
				csv_list = read_csv(p)
				exc_csv = extrude_csv(csv_list)
				ind = 0
				ind2 = 0
				for d in exc_csv:																		# Проверяем наличие номера
					try:
						if main_dt[d[0]]:
							fl = True
					except:
						fl = False
						n += 1
					if fl:																					# Создаем дикт для записи
						if int(d[1]):
							ind += 1																			
							fl3 = True
							report_dt[d[0]] = main_dt[d[0]]
							report_dt[d[0]]['Доставлено'] = main_dt[d[0]]['Доставлено'] + int(d[2])
							report_dt[d[0]]['Отправлено'] = main_dt[d[0]]['Отправлено'] + int(d[1])
							s = p.split('/')
							s = s[-1]
							if report_dt[d[0]]['Дата отправки'] == " ":
								report_dt[d[0]]['Дата отправки'] = s[:-4]
							else:
								report_dt[d[0]]['Дата отправки'] += ', ' +  s[:-4]
						else:
							ind2 +=1
				print('Отправлено ', ind, ' Не отправл ', ind2)
				print(p, ' не найдено ', n, ' номеров.')
			if report_dt:																					# Записываем дикты имеющие атрибуты
				os.chdir('провереные')
				rec_l = sort_dic_for_rec(report_dt)
				for d in rec_l:
					l_k = list(d.keys())
					path = d[l_k[0]]['Категории'] + '/' + d[l_k[0]]['Город'] + '.xlsx'
					flag = chek_file(path)
					if flag:
						rewrite_exele(path, d)
						print(path, ' дабавлено ', len(d),  ' отчетов.')
					else:
						print("Не удалось найти файл: ", path)
				back_dir(1)
			if fl3:
				for k in report_dt:
					main_dt[k] = report_dt[k]
				save_dict(main_dt)
				for p in p_l:
					os.rename('REPORTS/new/' + p , 'REPORTS/архив/' + p )
					print('перемещено REPORTS/архив/' + p)

def read_exel_in_dict(path):
	CHAR = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j' , 'k', 'l', 'm', 'n',
			'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'aa', 'ab',
			'ac', 'ad', 'ae', 'af', 'ag', 'ah', 'ai', 'aj', 'ak', 'al', 'am', 'an',
			'ao', 'ap', 'aq', 'ar', 'as']
	exel = openpyxl.load_workbook(path)
	sheet = exel.active
	cols = sheet.max_column
	rows = sheet.max_row
	l_key = []	
	for s in range(0, cols):
		k = CHAR[s] + '1'
		l_key.append(sheet[k].value)
	dt = {}
	l_n = []
	for r in sheet['a']:
		l_n.append(r.value)
	l_n = l_n[1:]
	i_ch = 1
	i_r = 1
	for n in l_n:
		dt[n] = {}
		i_r += 1
		for k in range(1, cols):
			dt[n][l_key[k]] = sheet[CHAR[k] + str(i_r)].value
	l = list(dt.keys())
	if 'Отправлено' not in dt[l[0]]:
		dt = value_assig(dt)
	return dt

def rewrite_exele(path, rec_dt):
	old_dt = read_exel_in_dict(path)
	l_key_old = get_keys(old_dt)
	l_val_main = [l_key_old]
	old_dt = remove_data(old_dt, rec_dt)
	old_dt.update(rec_dt)
	old_dt = sort_devilery(old_dt)
	l_val_main = dict_in_rec_list(old_dt, l_val_main)	
	save_exele(path, l_val_main)

def writin_new_exele(path, rec_dt):
	l_key = get_keys(rec_dt)
	l_val_main = [l_key]
	l_val_main = dict_in_rec_list(rec_dt, l_val_main)
	save_exele(path, l_val_main)

def chek_file(path):
	l = path.split("/")
	if len(l)>1:
		dir_ = "/".join(l[:-1])
	else:
		dir_ = None
	list_dir = os.listdir(dir_)
	flag = False
	for f in list_dir:
		if f == l[-1]:
			flag = True
			break
	return flag

def search_duble_in_dict(search_dt, main_dict):
	recdt = {}
	db = 0
	for sd in search_dt:
		fl = False
		for md in main_dict:
			if sd == md:
				fl = True
				db += 1
				break
		if not fl:
			recdt[sd] = search_dt[sd]
	return recdt, db

def back_dir(count_lev):
	pd = os.getcwd()
	l = pd.split('\\')
	l = l[:-count_lev]
	s = ''
	for i in l:
		s += i +'\\'
	os.chdir(s)

def init_dict():
	flag = chek_file('DICT.txt')
	if flag:
		with open("DICT.txt", 'rb') as DICT:
			dict_numb = pickle.load(DICT)
		print('Словарь загружен. ', len(dict_numb), ' номеров')
	else:
		if input('Словарь не найден, создать?: ') == 'y':
			print("Создается новый словарь")
			dict_numb = {}
			path_list = load_path_xl('провереные')
			for p in path_list:
				print(p)
				if "Старая база" in p:
			 		dt = read_exel_in_dict(p)
			 		li = p.split("/")
			 		sity = li[-1][:-5]
			 		cat = "/".join(li[:-1])
			 		for k in dt:
			 			dt[k]["Категории"] = cat
			 			dt[k]["Город"] = sity
			 		dict_numb.update(dt)
				else:
					dict_numb2 = read_exel_in_dict(p)
					dict_numb.update(dict_numb2)
			with open('DICT.txt', 'wb') as DICT_DEVELOR:
			 	pickle.dump(dict_numb, DICT_DEVELOR)
			print('Создалась новый словарь. ', len(dict_numb), ' номеров.')
	return dict_numb

def validate_category(dt):
	l = list(dt.keys())
	if 'Категории' in dt[l[0]]:
		for k in dt:
			s = dt[k]['Категории']
			i = 0
			for b in s:
				if b.isalpha():
					s = s[i:]
					break
				i += 1
			s = s.split('/')
			if s[0].strip() == 'Предложение услуг':
				s2 = 'Услуги/'
				for i in s:
					s2 += i.strip() + '/'
				dt[k]['Категории'] = s2.strip('/')
				continue
			if s[0].strip() == 'Услуги':
				s2 = ''
				for i in s:
					i = i.strip()
					s2 += i + '/'
				dt[k]['Категории'] = s2.strip('/')
			else:
				s2 = 'Услуги/Предложение услуг/'
				for i in s:
					i = i.strip()
					s2 += i + '/'
				dt[k]['Категории'] = s2.strip('/')
	return dt

def load_path_xl(folder):
	path = os.listdir(folder)
	path_list = []
	for n in path:
		if n[-1] =='x':
			p = folder + '/' + n
			path_list.append(p)
		else:
			p = folder + '/' + n
			l2 = os.listdir(p)
			for i in l2:
				if i[-1] =='x':
					p = folder + '/' + n + "/" + i
					path_list.append(p)
				else:
					p = folder + '/' + n + '/' + i
					l3 = os.listdir(p)
					for i2 in l3:
						if i2[-1] == "x":
							p = folder + '/' + n + '/' + i + '/' +i2
							path_list.append(p)
						else:
							p = folder + '/' + n + '/' + i + '/' +i2
							l4 = os.listdir(p)
							for i3 in l4:
								if i3[-1] =="x":
									p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3
									path_list.append(p)
								else:
									p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3
									l5 = os.listdir(p)
									for i4 in l5:
										if i4[-1] == 'x':
											p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 + '/' + i4
											path_list.append(p)
										else:
											p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 + '/' + i4
											l5 = os.listdir(p)
											for i5 in l5:
												if i5[-1] == 'x':
													p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 + '/' + i4 + '/' + i5
													path_list.append(p)
												else:
													p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 + '/' + i4 + '/' + i5
													l6 = os.listdir(p)
													for i6 in l6:
														if i6[-1] =='x':
															p = folder + '/' + n + '/' + i + '/' +i2 +'/' + i3 + '/' + i4 + '/' + i5 + '/' + i6
															path_list.append(p)
	
	return path_list

def main():
	main_dict = init_dict()
	init_report(main_dict)
	response = None
	while response != '0':
		print('''
			1 - Проверить файл
			2 - Колличество номеров в файлах
			3 - Поиск по номеру
			4 - Отчеты по смс
			5 - Валидация имен
			6 - Добавить в feedback
			10 - Проверка категорий
			0 - Выйти''')
		response = input("Введите: ")
		if response == '1':
			response1(main_dict)

		if response == '2':
			response2(main_dict)

		if response == '3':
			response3(main_dict)

		if response =='4':
			response4(main_dict)
		
		if response == '5':
			import name
			name.main()

		if response =='6':
			add_feedback()

		if response =='10':
			response10(main_dict)
		
		if response == '7':
			dt = search_bulding_person(main_dict)
			print(len(dt))
			dt_deliv = {}
			for k in dt:
				for w in SITY_VAL:
					if w.lower() == dt[k]['Город'].lower():
						if dt[k]['Отправлено'] == 0 or dt[k]['Отправлено'] == 1 and 'натяжн' not in dt[k]['Заголовок'].lower():
							dt_deliv[k] = dt[k]
							break	
			print(len(dt_deliv))
			L_sort =[]
			for k in dt_deliv:
				if dt_deliv[k]['Дата отправки'] == " ":
					L_sort.append((k, 0))
				else:
					c = dt_deliv[k]['Дата отправки'].split('.')
					c.reverse()
					c = ''.join(c)
					L_sort.append((k, int(c)))
			L_sort.sort(key=lambda x: x[1])
			dt = {}
			for c in L_sort:
				dt[c[0]] = main_dict[c[0]]
			writin_new_exele('333.xlsx', dt)

		if response == '8':
			dt = validate_sity(main_dict)
			dt = validate_celling(dt)
			targ = {}
			for k in dt:
				if dt[k]['Отправлено'] == 0:
					targ[k] =dt[k]
			writin_new_exele('123.xlsx', targ)
		if response == '11':
			import test
		if response == '12':
			test.foo()

			
	
main()